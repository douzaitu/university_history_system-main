from DrissionPage import ChromiumPage, ChromiumOptions
import time
import random
import json
import os
import re
import requests
import urllib.parse
from loguru import logger
from bs4 import BeautifulSoup
from openpyxl import Workbook
from requests.exceptions import RequestException
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

class TeacherCrawler:
    def __init__(self):
        # 1. 配置 Chrome 路径（关键！替换为你的实际路径）
        chrome_options = ChromiumOptions()
        # 替换为你的 Chrome 可执行文件路径（示例路径，需修改！）
        chrome_options.set_browser_path(r"C:\Users\ASUS\AppData\Local\Google\Chrome\Application\chrome.exe")

        # 初始化浏览器（自动处理SSL证书、请求头）
        self.browser = ChromiumPage(chrome_options)  # 修复：之前漏传chrome_options参数
        self.browser.set.window.max()  # 最大化窗口，模拟真实用户
        self.list_url = "https://cist.cdut.edu.cn/szdw/jsml.htm"  # 目标URL
        self.save_dir = r"D:\spider_project\data\raw\teacher_page"  # 数据保存路径
        self.data_list = []  # 用于存储所有导师信息，供Excel保存

        # 反检测JS代码（绕过瑞数等JS检测）
        self.anti_detection_js = """
        // 禁用debugger
        window.debugger = function(){};
        Object.defineProperty(window, 'debugger', {configurable: false, writable: false});

        // 覆盖自动化检测属性
        Object.defineProperty(navigator, 'webdriver', {get: () => false});
        Object.defineProperty(window, 'navigator', {configurable: false, writable: false});

        // 删除浏览器自动化特征
        delete window.chrome;
        """

    def simulate_human(self):
        """模拟人类操作行为，降低反爬识别概率"""
        try:
            # 直接定位body，超时5秒，找不到则跳过
            body = self.browser.ele("tag:body", timeout=3)
            if body:
                # 随机滚动
                for _ in range(random.randint(2, 4)):
                    scroll_px = random.randint(200, 600)
                    self.browser.scroll.down(scroll_px)
                    time.sleep(random.uniform(0.3, 1.0))
                body.click()
                time.sleep(random.uniform(0.5, 1.0))
            else:
                print("未找到页面body，跳过模拟操作")
        except Exception as e:
            print(f"模拟人类操作失败：{str(e)}")
            # 跳过错误，继续爬虫流程

    def bypass_ruishi(self, url):
        """处理瑞数安全验证"""
        try:
            self.browser.get(url)
            self.browser.run_js(self.anti_detection_js)
            time.sleep(3)  #
            print("页面访问完成，自动继续")
            return True
        except Exception as e:
            print(f"页面访问失败：{str(e)}，跳过当前页面")
            return False

    def get_teacher_links(self):
        title_links = [
            "https://cist.cdut.edu.cn/info/1121/6384.htm",
            "https://cist.cdut.edu.cn/info/1121/2538.htm",
            "https://cist.cdut.edu.cn/info/1121/2370.htm"
        ]
        print(f"直接访问3个职称页：{title_links}")

        # 只爬这3个页面里的导师链接（info/1121目录下）
        teacher_links = []
        allowed_dirs = ["1121", "1118", "1045"]
        seen_links = set()

        for url in title_links:
            print(f"\n正在访问：{url}")
            self.bypass_ruishi(url)
            self.simulate_human()
            time.sleep(5)
            soup = BeautifulSoup(self.browser.html, "html.parser")

            # 只提该页面下 info/1121 目录的导师链接（姓名2-4字）
            for a in soup.find_all("a"):
                name = a.get_text(strip=True).replace(" ", "")
                href = a.get("href")

                if (len(name) >= 2 and len(name) <= 4 and
                        href and href.endswith(".htm") and
                        # all('\u4e00' <= char <= '\u9fa5' for char in name)
                        any(dir_str in href for dir_str in allowed_dirs)):
                    full_href = ""
                    if "info/" in href:
                        # 已有 info/，直接拼域名（比如 href 是 "/info/1118/2382.htm"）
                        full_href = f"https://cist.cdut.edu.cn{href.lstrip('/')}"
                    else:
                        # 修复次要错误2：target_dir 容错，避免索引错误
                        target_dir_list = [d for d in allowed_dirs if d in href]
                        if not target_dir_list:  # 没找到目录时跳过，不报错
                            continue
                        # 没有 info/，提取目录+编号拼接（比如 href 是 "1118/2382.htm" 或 "2382.htm"）
                        target_dir = target_dir_list[0]  # 提取1118/1121/1045
                        file_name = href.split("/")[-1]  # 提取编号.htm（比如 2382.htm）
                        full_href = f"https://cist.cdut.edu.cn/info/{target_dir}/{file_name}"

        # 有序去重：没见过的链接才添加，保留原始顺序
                    if full_href not in seen_links:
                        seen_links.add(full_href)
                        teacher_links.append(full_href)
                        print(f"有效链接（顺序{len(teacher_links)}）：{full_href}")

        # teacher_links = list(set(teacher_links))
        print(f"\n共提取到 {len(teacher_links)} 个导师链接")
        return teacher_links


    def parse_teacher_detail(self, html):
        soup = BeautifulSoup(html, "html.parser")

        # 1. 提取姓名（从页面标题提取）
        name = "未知"
        if soup.title:
            title_text = soup.title.get_text(strip=True)
            if "-" in title_text:
                name = title_text.split("-")[0].strip()

        # 2. 提取简介（从 meta description 提取）
        intro = "未知"
        desc_meta = soup.find('meta', attrs={'name': 'description'})
        if desc_meta and 'content' in desc_meta.attrs:
            intro = desc_meta['content'].strip()

        # 3. 提取正文详细内容（从 #vsb_content 下的 p 标签提取）
        full_content = "未知"
        content_div = soup.select_one('#vsb_content .v_news_content')
        if content_div:
            paragraphs = content_div.find_all('p')
            full_content = '\n'.join([p.get_text(strip=True) for p in paragraphs if p.get_text(strip=True)])

        # 4.图片爬取
        img_path = "无"  # 初始化默认值
        try:
            img_tag = content_div.find('img', src=True) if content_div else None
            if not img_tag:
                print(f"  {name} 无图片可爬取")
            else:
                #处理图片URL（相对路径转绝对路径）
                img_url = img_tag['src'].strip()
                if not img_url.startswith(('http', 'https')):
                    img_url = urllib.parse.urljoin("https://cist.cdut.edu.cn/", img_url)
                print(f"  正在下载图片：{img_url}")
                #下载图片（避免反爬）
                headers = {
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
                }
                max_retries = 3  # 最多重试3次（解决网络波动问题）
                #保存图片
                img_dir = os.path.join(self.save_dir, "images")  # 图片存放在 save_dir/images 下
                os.makedirs(img_dir, exist_ok=True)  # 自动创建目录（无需手动建）
                img_path = os.path.join(img_dir, f"{name}.jpg")  # 文件名：姓名.jpg

                #下载并保存
                response = requests.get(img_url, timeout=10)
                with open(img_path, 'wb') as f:
                    f.write(response.content)
                    print(f"{name} 图片已保存至：{img_path}")
        except Exception as e:
            img_path = "无"
            print(f"{name} 图片处理失败：{str(e)}")

        # 保存详情页HTML到绝对路径
        html_dir = self.save_dir  # HTML保存到teacher_page目录下
        os.makedirs(html_dir, exist_ok=True)
        html_path = os.path.join(html_dir, f"{name}.html")
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html)  # html是parse_teacher_detail的参数，直接写入
        print(f"{name} 详情页已保存至：{html_path}")

        # 5. 收集数据
        html_path = os.path.join(self.save_dir, f"{name}.html")  # 定义HTML路径
        self.data_list.append({
            "姓名": name,
            "简介": intro,
            "详细内容": full_content,
            "html_path": html_path,
            "img_path": img_path
        })


        # 输出提取结果
        print(f"\n===== 导师信息提取结果 =====")
        print(f"姓名：{name}")
        print(f"简介：{intro}")
        print(f"详细内容：\n{full_content}")
        print("-" * 50)

    def crawl(self):
        """核心爬取流程："""
        print("===== 启动爬虫 =====")
        # 1. 提链接
        links = self.get_teacher_links()
        print(f"共提{len(links)}个导师链接")
        # 2. 爬详情页
        try:
            for i, link in enumerate(links, 1):
                print(f"\n爬第{i}个导师：{link}")
                try:
                    self.bypass_ruishi(link)
                    time.sleep(3)
                    self.parse_teacher_detail(self.browser.html)
                except Exception as e:
                    print(f"爬取失败：{e}，跳过")

        finally:
            self.browser.quit()
            self.save_to_excel(self.data_list)  # 仅最终保存一次Excel
        print("===== 爬完收工 =====")

    def save_to_excel(self, data_list):
        """将所有导师信息保存到Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "导师信息"
        # 写入表头
        ws.append(["姓名", "简介", "详细内容", "详情页HTML路径", "图片"])

        # 写入每条数据
        for row_idx, data in enumerate(data_list, 2):
            name = data["姓名"]
            intro = data["简介"]
            full_content = data["详细内容"]
            html_path = data["html_path"]
            img_path = data.get("img_path", "无")


            # 写入文本数据
            ws.append([
                name,
                intro,
                full_content,
                html_path,
                ""
            ])

            # 插入图片
            if img_path != "无" and os.path.exists(img_path):
                try:
                    img = Image(img_path)
                    img.width = 120
                    img.height = 150
                    col = get_column_letter(5)
                    ws.add_image(img, f"{col}{row_idx}")
                    ws.column_dimensions[col].width = 18
                    ws.row_dimensions[row_idx].height = 120
                except Exception as e:
                    print(f"插入{name}图片失败：{str(e)}")

        # 优化其他列宽（提升显示效果）
        ws.column_dimensions["A"].width = 10  # 姓名
        ws.column_dimensions["B"].width = 30  # 简介
        ws.column_dimensions["C"].width = 50  # 详细内容
        ws.column_dimensions["D"].width = 40  # 详情页HTML路径

        # 保存Excel
        excel_path = os.path.join("D:\\spider_project\\data\\processed", "导师信息汇总.xlsx")
        os.makedirs(os.path.dirname(excel_path), exist_ok=True)
        wb.save(excel_path)
        print(f"Excel已保存至：{excel_path}")

